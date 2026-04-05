"""
tests/desktop/test_export_endpoints.py

Tests the PDF and Excel generation logic directly — no HTTP stack, no database needed.

Run from the backend directory:
    cd backend
    python -m pytest tests/desktop/test_export_endpoints.py -v
"""
import sys
import os
import io

# Ensure backend/ is on sys.path
_BACKEND = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
if _BACKEND not in sys.path:
    sys.path.insert(0, _BACKEND)

import pytest


# ─────────────────────────────────────────────────────────────────────────────
# PDF Generation Tests  (utils/pdf_generator.py — already uses reportlab only)
# ─────────────────────────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def pdf_gen():
    from utils.pdf_generator import generate_report_pdf
    return generate_report_pdf


def _voucher_data():
    return {
        "total_amount": 100000.0,
        "total_count": 2,
        "tax_estimate": 18000.0,
        "vouchers": [
            {
                "id": 1, "date": "01-Apr-2025", "date_iso": "2025-04-01",
                "party_name": "Test Party", "voucher_type": "Sales",
                "voucher_no": "S/001", "amount": 50000.0, "narration": "Sample",
            },
            {
                "id": 2, "date": "02-Apr-2025", "date_iso": "2025-04-02",
                "party_name": "Another Party", "voucher_type": "Sales",
                "voucher_no": "S/002", "amount": 50000.0, "narration": None,
            },
        ],
    }


def test_sales_register_pdf_non_empty(pdf_gen):
    result = pdf_gen(
        slug="sales-register",
        report_title="Sales Register",
        data=_voucher_data(),
        period="01 Apr 2025 – 31 Mar 2026",
    )
    assert isinstance(result, bytes)
    assert len(result) > 100
    assert result[:4] == b"%PDF", f"Expected PDF header, got: {result[:8]!r}"


def test_balance_sheet_pdf_non_empty(pdf_gen):
    data = {
        "assets":           [{"name": "Cash", "amount": 50000, "group": "Current Assets"}],
        "liabilities":      [{"name": "Creditors", "amount": 30000, "group": "Current Liabilities"}],
        "total_assets":     50000,
        "total_liabilities": 30000,
        "net_difference":   20000,
    }
    result = pdf_gen(
        slug="balance-sheet",
        report_title="Balance Sheet",
        data=data,
        period="01 Apr 2025 – 31 Mar 2026",
    )
    assert result[:4] == b"%PDF"
    assert len(result) > 100


def test_profit_loss_pdf_non_empty(pdf_gen):
    data = {
        "income":         {"Sales Accounts": 200000.0, "Direct Income": 0.0},
        "expenses":       {"Purchase Accounts": 150000.0, "Direct Expenses": 0.0},
        "total_income":   200000.0,
        "total_expenses": 150000.0,
        "net_profit":     50000.0,
    }
    result = pdf_gen(
        slug="profit-loss",
        report_title="Profit & Loss",
        data=data,
        period="01 Apr 2025 – 31 Mar 2026",
    )
    assert result[:4] == b"%PDF"
    assert len(result) > 100


# ─────────────────────────────────────────────────────────────────────────────
# Excel Generation Tests  (utils/excel_generator.py — openpyxl only)
# ─────────────────────────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def excel_gen():
    from utils.excel_generator import generate_report_excel
    return generate_report_excel


def _open_xlsx(raw: bytes):
    import openpyxl
    return openpyxl.load_workbook(io.BytesIO(raw))


def test_sales_register_excel_non_empty(excel_gen):
    data = {
        "vouchers": [
            {
                "date": "01-Apr-2025", "party_name": "Test Party",
                "voucher_type": "Sales", "voucher_no": "S/001",
                "narration": "Sample", "amount": 50000.0,
            }
        ]
    }
    raw = excel_gen("sales-register", "Sales Register", data, "01 Apr 2025 – 31 Mar 2026")
    assert isinstance(raw, bytes)
    assert len(raw) > 100
    wb = _open_xlsx(raw)
    ws = wb.active
    assert ws.max_row >= 5, f"Expected >=5 rows, got {ws.max_row}"


def test_balance_sheet_excel_non_empty(excel_gen):
    data = {
        "assets":      [{"name": "Cash",      "amount": 50000, "group": "Current Assets"}],
        "liabilities": [{"name": "Creditors", "amount": 30000, "group": "Current Liabilities"}],
    }
    raw = excel_gen("balance-sheet", "Balance Sheet", data, "01 Apr 2025 – 31 Mar 2026")
    assert len(raw) > 100
    wb = _open_xlsx(raw)
    ws = wb.active
    assert ws.max_row >= 7


def test_profit_loss_excel_non_empty(excel_gen):
    data = {
        "income":   {"Sales Accounts": 200000.0},
        "expenses": {"Purchase Accounts": 150000.0},
    }
    raw = excel_gen("profit-loss", "Profit & Loss", data, "01 Apr 2025 – 31 Mar 2026")
    assert len(raw) > 100
    wb = _open_xlsx(raw)
    ws = wb.active
    assert ws.max_row >= 6


def test_excel_header_row_has_correct_columns(excel_gen):
    raw = excel_gen("sales-register", "Sales Register", {"vouchers": []}, "some period")
    wb  = _open_xlsx(raw)
    ws  = wb.active
    # Row 5 = header (rows 1-3 meta + row 4 blank)
    header_row = [ws.cell(row=5, column=c).value for c in range(1, 8)]
    assert "#" in header_row
    assert "Date" in header_row
    assert "Amount (INR)" in header_row


def test_unknown_slug_falls_back_to_placeholder(excel_gen):
    raw = excel_gen("gst-summary", "GST Summary", {}, "some period")
    assert len(raw) > 100
    wb = _open_xlsx(raw)
    ws = wb.active
    all_values = [
        ws.cell(row=r, column=c).value
        for r in range(1, ws.max_row + 1)
        for c in range(1, ws.max_column + 1)
    ]
    assert any(v and "PLACEHOLDER" in str(v) for v in all_values)
