"""
Export Service - PDF & Excel Generation for Desktop App
========================================================
Generates professional PDF invoices, statements, and Excel reports
stored locally on the user's machine for WhatsApp delivery.

Desktop App Flow:
1. User requests PDF/Excel via WhatsApp
2. Backend generates file in local exports folder
3. Baileys listener sends file via WhatsApp
4. File is also viewable in K24 Dashboard

Key Features:
- Professional PDF invoices matching Tally format
- Excel exports with formatting
- Outstanding statements
- Sales/Purchase registers
"""

import os
import io
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path

from sqlalchemy.orm import Session
from sqlalchemy import func, desc

# PDF Generation
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib import colors
from reportlab.lib.units import inch, mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

# Excel Generation
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from backend.database import (
    SessionLocal, Ledger, Voucher, StockItem, Bill, StockMovement
)

logger = logging.getLogger("ExportService")

# Get exports directory (in app data folder for desktop)
def get_exports_dir() -> Path:
    """Get the exports directory, creating if needed"""
    # Use the app's data folder (works for both dev and packaged desktop app)
    base_dir = Path(os.environ.get("K24_DATA_DIR", os.path.dirname(os.path.dirname(__file__))))
    exports_dir = base_dir / "exports"
    exports_dir.mkdir(parents=True, exist_ok=True)
    return exports_dir


class PDFGenerator:
    """Generates professional PDF documents"""
    
    def __init__(self, db: Session, tenant_id: str = "default"):
        self.db = db
        self.tenant_id = tenant_id
        self.exports_dir = get_exports_dir()
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Setup custom paragraph styles"""
        self.styles.add(ParagraphStyle(
            name='CompanyName',
            parent=self.styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor("#1a365d"),
            alignment=TA_CENTER,
            spaceAfter=6
        ))
        self.styles.add(ParagraphStyle(
            name='DocumentTitle',
            parent=self.styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor("#2d3748"),
            alignment=TA_CENTER,
            spaceAfter=12
        ))
        self.styles.add(ParagraphStyle(
            name='TableHeader',
            parent=self.styles['Normal'],
            fontSize=10,
            textColor=colors.white,
            alignment=TA_CENTER
        ))
        self.styles.add(ParagraphStyle(
            name='RightAlign',
            parent=self.styles['Normal'],
            alignment=TA_RIGHT
        ))
    
    def generate_invoice_pdf(self, voucher_id: int) -> Tuple[str, str]:
        """
        Generate a professional GST Tax Invoice PDF matching Tally Prime format.
        
        Features:
        - Company header with GSTIN
        - Buyer and Consignee details
        - HSN/SAC codes for items
        - Tax breakdowns (CGST, SGST, IGST)
        - Amount in words
        - Authorized signatory
        
        Returns:
            Tuple of (file_path, filename)
        """
        # Fetch voucher data
        voucher = self.db.query(Voucher).filter(
            Voucher.id == voucher_id,
            Voucher.tenant_id == self.tenant_id
        ).first()
        
        if not voucher:
            raise ValueError(f"Voucher {voucher_id} not found")
        
        # Fetch party details
        party = None
        if voucher.party_name:
            party = self.db.query(Ledger).filter(
                Ledger.name == voucher.party_name,
                Ledger.tenant_id == self.tenant_id
            ).first()
        
        # Fetch stock movements for this voucher (line items)
        items = self.db.query(StockMovement).filter(
            StockMovement.voucher_id == voucher_id
        ).all()
        
        # Generate filename
        voucher_num = voucher.voucher_number or f"V{voucher_id}"
        safe_name = "".join(c for c in voucher_num if c.isalnum() or c in "-_")
        filename = f"Invoice_{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = self.exports_dir / filename
        
        # Create PDF
        doc = SimpleDocTemplate(
            str(filepath),
            pagesize=A4,
            rightMargin=15*mm,
            leftMargin=15*mm,
            topMargin=15*mm,
            bottomMargin=15*mm
        )
        
        elements = []
        
        # ============== COMPANY HEADER ==============
        elements.append(Paragraph(
            "<b>TAX INVOICE</b>",
            ParagraphStyle('Title', fontSize=14, alignment=TA_CENTER, textColor=colors.HexColor("#1a365d"))
        ))
        elements.append(Spacer(1, 8))
        
        # Company Info Box
        company_info = [
            ["<b>KRISHA SALES</b>", "", "<b>Invoice Details</b>"],
            ["Address: [Your Business Address]", "", f"Invoice No: <b>{voucher_num}</b>"],
            ["State: Rajasthan  |  Code: 08", "", f"Date: {voucher.date.strftime('%d-%b-%Y') if voucher.date else 'N/A'}"],
            ["GSTIN: 08AAAAA0000A1Z5", "", f"Type: {voucher.voucher_type or 'Sales'}"],
            ["E-mail: sales@krishasales.com", "", "E-Way Bill: -"],
        ]
        
        company_table = Table(company_info, colWidths=[200, 50, 200])
        company_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('LINEAFTER', (0, 0), (0, -1), 0.5, colors.grey),
            ('LINEAFTER', (-2, 0), (-2, -1), 0.5, colors.grey),
        ]))
        elements.append(company_table)
        elements.append(Spacer(1, 10))
        
        # ============== BUYER / CONSIGNEE DETAILS ==============
        buyer_name = voucher.party_name or "Cash Sale"
        buyer_gstin = party.gstin if party else "-"
        buyer_address = party.address if party else ""
        buyer_state = party.state if party else "Rajasthan"
        
        buyer_info = [
            ["<b>Bill To:</b>", "<b>Ship To (Consignee):</b>"],
            [f"Name: {buyer_name}", f"Name: {buyer_name}"],
            [f"GSTIN: {buyer_gstin}", f"Address: {buyer_address or '-'}"],
            [f"State: {buyer_state}  |  Code: 08", f"State: {buyer_state}"],
        ]
        
        buyer_table = Table(buyer_info, colWidths=[225, 225])
        buyer_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('LINEAFTER', (0, 0), (0, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#f0f0f0")),
        ]))
        elements.append(buyer_table)
        elements.append(Spacer(1, 10))
        
        # ============== ITEMS TABLE ==============
        items_header = [["Sl", "Description of Goods", "HSN", "Qty", "Unit", "Rate", "Amount"]]
        items_data = []
        
        subtotal = 0.0
        for idx, item in enumerate(items, 1):
            # Get item details
            stock_item = self.db.query(StockItem).filter(
                StockItem.id == item.item_id,
                StockItem.tenant_id == self.tenant_id
            ).first()
            
            item_name = stock_item.name if stock_item else f"Item {item.item_id}"
            hsn_code = stock_item.hsn_code if stock_item else "-"
            unit = stock_item.units if stock_item else "Nos"
            
            qty = item.quantity or 0
            rate = item.rate or 0
            amount = qty * rate
            subtotal += amount
            
            items_data.append([
                str(idx),
                Paragraph(item_name, self.styles['Normal']),
                hsn_code or "-",
                f"{qty:.2f}",
                unit,
                f"₹{rate:,.2f}",
                f"₹{amount:,.2f}"
            ])
        
        # If no items, show the voucher total as a single line
        if not items_data:
            subtotal = voucher.amount or 0
            items_data.append(["1", "As per invoice", "-", "1", "Nos", f"₹{subtotal:,.2f}", f"₹{subtotal:,.2f}"])
        
        # Create items table
        all_items = items_header + items_data
        items_table = Table(all_items, colWidths=[25, 180, 50, 40, 35, 60, 70])
        items_table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1a365d")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            # Body
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (2, 1), (2, -1), 'CENTER'),
            ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(items_table)
        
        # ============== TAX SUMMARY ==============
        # Calculate GST (assuming 18% for now, can be made dynamic)
        gst_rate = 18.0
        cgst = subtotal * (gst_rate / 2) / 100
        sgst = subtotal * (gst_rate / 2) / 100
        total = subtotal + cgst + sgst
        
        # Override with actual voucher amount if available
        if voucher.amount:
            total = voucher.amount
        
        tax_data = [
            ["", "", "", "", "Taxable Value:", f"₹{subtotal:,.2f}"],
            ["", "", "", "", f"CGST @ {gst_rate/2}%:", f"₹{cgst:,.2f}"],
            ["", "", "", "", f"SGST @ {gst_rate/2}%:", f"₹{sgst:,.2f}"],
            ["", "", "", "", "<b>TOTAL:</b>", f"<b>₹{total:,.2f}</b>"],
        ]
        
        tax_table = Table(tax_data, colWidths=[25, 180, 50, 40, 85, 80])
        tax_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (-2, 0), (-1, -1), 'RIGHT'),
            ('LINEABOVE', (-2, -1), (-1, -1), 1, colors.black),
        ]))
        elements.append(tax_table)
        elements.append(Spacer(1, 10))
        
        # ============== AMOUNT IN WORDS ==============
        amount_words = self._number_to_words(total)
        elements.append(Paragraph(
            f"<b>Amount in Words:</b> {amount_words} Only",
            self.styles['Normal']
        ))
        elements.append(Spacer(1, 20))
        
        # ============== FOOTER / SIGNATURE ==============
        footer_data = [
            ["<b>Terms & Conditions:</b>", "", "<b>For KRISHA SALES</b>"],
            ["1. Goods once sold will not be taken back.", "", ""],
            ["2. Subject to Jaipur jurisdiction only.", "", ""],
            ["", "", "Authorised Signatory"],
        ]
        
        footer_table = Table(footer_data, colWidths=[250, 50, 150])
        footer_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (-1, 0), (-1, -1), 'CENTER'),
        ]))
        elements.append(footer_table)
        
        # K24 watermark
        elements.append(Spacer(1, 10))
        elements.append(Paragraph(
            f"<i>Generated by K24.ai on {datetime.now().strftime('%d-%b-%Y %H:%M')}</i>",
            ParagraphStyle('Footer', fontSize=7, textColor=colors.grey, alignment=TA_CENTER)
        ))
        
        # Build PDF
        doc.build(elements)
        
        logger.info(f"Generated Tally-style invoice PDF: {filepath}")
        return str(filepath), filename
    
    def _number_to_words(self, num: float) -> str:
        """Convert number to Indian currency words"""
        ones = ['', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight', 'Nine', 
                'Ten', 'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen', 'Sixteen', 
                'Seventeen', 'Eighteen', 'Nineteen']
        tens = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty', 'Sixty', 'Seventy', 'Eighty', 'Ninety']
        
        def convert_chunk(n):
            if n < 20:
                return ones[n]
            elif n < 100:
                return tens[n // 10] + (' ' + ones[n % 10] if n % 10 else '')
            else:
                return ones[n // 100] + ' Hundred' + (' ' + convert_chunk(n % 100) if n % 100 else '')
        
        if num == 0:
            return 'Zero Rupees'
        
        num = int(round(num))
        
        crores = num // 10000000
        num %= 10000000
        lakhs = num // 100000
        num %= 100000
        thousands = num // 1000
        num %= 1000
        hundreds = num
        
        result = ''
        if crores:
            result += convert_chunk(crores) + ' Crore '
        if lakhs:
            result += convert_chunk(lakhs) + ' Lakh '
        if thousands:
            result += convert_chunk(thousands) + ' Thousand '
        if hundreds:
            result += convert_chunk(hundreds)
        
        return 'Rupees ' + result.strip()
    
    def generate_outstanding_statement(self, party_name: str) -> Tuple[str, str]:
        """
        Generate outstanding statement PDF for a party.
        
        Returns:
            Tuple of (file_path, filename)
        """
        # Find ledger
        ledger = self.db.query(Ledger).filter(
            Ledger.name.ilike(f"%{party_name}%"),
            Ledger.tenant_id == self.tenant_id
        ).first()
        
        if not ledger:
            raise ValueError(f"Party '{party_name}' not found")
        
        # Get outstanding bills - use correct field names from database schema
        # Bill has: bill_name, party_name, amount, due_date, is_overdue
        bills = self.db.query(Bill).filter(
            Bill.party_name.ilike(f"%{ledger.name}%"),
            Bill.amount > 0,  # Positive = Receivable
            Bill.tenant_id == self.tenant_id
        ).order_by(Bill.due_date).all()
        
        # Generate filename
        safe_name = "".join(c for c in ledger.name if c.isalnum() or c in "-_ ")[:30]
        filename = f"Statement_{safe_name}_{datetime.now().strftime('%Y%m%d')}.pdf"
        filepath = self.exports_dir / filename
        
        # Create PDF
        doc = SimpleDocTemplate(
            str(filepath),
            pagesize=A4,
            rightMargin=20*mm,
            leftMargin=20*mm,
            topMargin=20*mm,
            bottomMargin=20*mm
        )
        
        elements = []
        
        # Header
        elements.append(Paragraph("KRISHA SALES", self.styles['CompanyName']))
        elements.append(Paragraph("Outstanding Statement", self.styles['DocumentTitle']))
        elements.append(Spacer(1, 10))
        
        # Party Info with GSTIN if available
        elements.append(Paragraph(f"<b>Party:</b> {ledger.name}", self.styles['Normal']))
        if ledger.gstin:
            elements.append(Paragraph(f"<b>GSTIN:</b> {ledger.gstin}", self.styles['Normal']))
        elements.append(Paragraph(f"<b>As on:</b> {datetime.now().strftime('%d-%b-%Y')}", self.styles['Normal']))
        elements.append(Spacer(1, 15))
        
        # Calculate totals - amount field is the outstanding amount
        total_outstanding = sum(abs(b.amount or 0) for b in bills)
        total_overdue = sum(abs(b.amount or 0) for b in bills if b.is_overdue or (b.due_date and b.due_date < datetime.now()))
        
        # Bills Table
        bills_data = [["#", "Bill Ref", "Due Date", "Amount", "Status"]]
        
        for idx, bill in enumerate(bills, 1):
            is_overdue = bill.is_overdue or (bill.due_date and bill.due_date < datetime.now())
            status = "Overdue" if is_overdue else "Current"
            
            bills_data.append([
                str(idx),
                bill.bill_name or "Bill",
                bill.due_date.strftime("%d-%b-%Y") if bill.due_date else "N/A",
                f"₹{abs(bill.amount or 0):,.2f}",
                status
            ])
        
        # If no bills, show ledger closing balance
        if not bills:
            closing = ledger.closing_balance or 0
            if closing > 0:
                bills_data.append(["1", "As per books", "-", f"₹{closing:,.2f}", "Current"])
                total_outstanding = closing
        
        # Summary row
        bills_data.append(["", "", "TOTAL:", f"₹{total_outstanding:,.2f}", ""])
        
        bills_table = Table(bills_data, colWidths=[30, 150, 90, 100, 80])
        bills_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1a365d")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
            ('FONTNAME', (2, -1), (3, -1), 'Helvetica-Bold'),
        ]))
        elements.append(bills_table)
        elements.append(Spacer(1, 20))
        
        # Summary Box
        summary_data = [
            ["Summary"],
            [f"Total Outstanding: ₹{total_outstanding:,.2f}"],
            [f"Overdue Amount: ₹{total_overdue:,.2f}"],
            [f"Current Due: ₹{total_outstanding - total_overdue:,.2f}"],
        ]
        
        summary_table = Table(summary_data, colWidths=[200])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOX', (0, 0), (-1, -1), 1, colors.grey),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 30))
        
        # Footer
        elements.append(Paragraph(
            f"<i>Generated by K24.ai on {datetime.now().strftime('%d-%b-%Y %H:%M')}</i>",
            ParagraphStyle('Footer', fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
        ))
        
        doc.build(elements)
        
        logger.info(f"Generated statement PDF: {filepath}")
        return str(filepath), filename


class ExcelGenerator:
    """Generates formatted Excel reports"""
    
    def __init__(self, db: Session, tenant_id: str = "default"):
        self.db = db
        self.tenant_id = tenant_id
        self.exports_dir = get_exports_dir()
        
        # Define styles
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_sales_register(
        self, 
        date_from: datetime = None, 
        date_to: datetime = None
    ) -> Tuple[str, str]:
        """
        Generate Sales Register Excel report.
        
        Returns:
            Tuple of (file_path, filename)
        """
        if not date_from:
            date_from = datetime.now() - timedelta(days=30)
        if not date_to:
            date_to = datetime.now()
        
        # Query sales vouchers
        vouchers = self.db.query(Voucher).filter(
            Voucher.voucher_type == "Sales",
            Voucher.date >= date_from,
            Voucher.date <= date_to,
            Voucher.tenant_id == self.tenant_id
        ).order_by(Voucher.date).all()
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Register"
        
        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = f"Sales Register ({date_from.strftime('%d-%b-%Y')} to {date_to.strftime('%d-%b-%Y')})"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Company name
        ws['A2'] = "KRISHA SALES"
        ws['A2'].font = Font(bold=True, size=10, color="666666")
        
        # Headers - simplified without VoucherItem dependency
        headers = ["Date", "Voucher No", "Party Name", "Amount", "Narration"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Data rows
        total_amount = 0.0
        
        for row_idx, voucher in enumerate(vouchers, 5):
            ws.cell(row=row_idx, column=1, value=voucher.date.strftime("%d-%b-%Y") if voucher.date else "")
            ws.cell(row=row_idx, column=2, value=voucher.voucher_number or "")
            ws.cell(row=row_idx, column=3, value=voucher.party_name or "Cash")
            
            amount = voucher.amount or 0
            ws.cell(row=row_idx, column=4, value=amount).number_format = '₹#,##0.00'
            ws.cell(row=row_idx, column=5, value=voucher.narration or "")
            
            total_amount += amount
            
            # Apply borders
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).border = self.border
        
        # Total row
        total_row = len(vouchers) + 5
        ws.cell(row=total_row, column=3, value="TOTAL:").font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=total_amount).number_format = '₹#,##0.00'
        ws.cell(row=total_row, column=4).font = Font(bold=True)
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30
        
        # Save
        filename = f"Sales_Register_{date_from.strftime('%Y%m%d')}_to_{date_to.strftime('%Y%m%d')}.xlsx"
        filepath = self.exports_dir / filename
        wb.save(filepath)
        
        logger.info(f"Generated sales register Excel: {filepath}")
        return str(filepath), filename
    
    def generate_purchase_register(
        self, 
        date_from: datetime = None, 
        date_to: datetime = None
    ) -> Tuple[str, str]:
        """Generate Purchase Register Excel report."""
        if not date_from:
            date_from = datetime.now() - timedelta(days=30)
        if not date_to:
            date_to = datetime.now()
        
        vouchers = self.db.query(Voucher).filter(
            Voucher.voucher_type == "Purchase",
            Voucher.date >= date_from,
            Voucher.date <= date_to,
            Voucher.tenant_id == self.tenant_id
        ).order_by(Voucher.date).all()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Purchase Register"
        
        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = f"Purchase Register ({date_from.strftime('%d-%b-%Y')} to {date_to.strftime('%d-%b-%Y')})"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Company name
        ws['A2'] = "KRISHA SALES"
        ws['A2'].font = Font(bold=True, size=10, color="666666")
        
        # Headers - simplified
        headers = ["Date", "Voucher No", "Supplier Name", "Amount", "Narration"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Data
        total_amount = 0.0
        for row_idx, voucher in enumerate(vouchers, 5):
            ws.cell(row=row_idx, column=1, value=voucher.date.strftime("%d-%b-%Y") if voucher.date else "")
            ws.cell(row=row_idx, column=2, value=voucher.voucher_number or "")
            ws.cell(row=row_idx, column=3, value=voucher.party_name or "Cash")
            
            amount = voucher.amount or 0
            ws.cell(row=row_idx, column=4, value=amount).number_format = '₹#,##0.00'
            ws.cell(row=row_idx, column=5, value=voucher.narration or "")
            
            total_amount += amount
            
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).border = self.border
        
        # Total
        total_row = len(vouchers) + 5
        ws.cell(row=total_row, column=3, value="TOTAL:").font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=total_amount).number_format = '₹#,##0.00'
        ws.cell(row=total_row, column=4).font = Font(bold=True)
        
        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30
        
        filename = f"Purchase_Register_{date_from.strftime('%Y%m%d')}_to_{date_to.strftime('%Y%m%d')}.xlsx"
        filepath = self.exports_dir / filename
        wb.save(filepath)
        
        logger.info(f"Generated purchase register Excel: {filepath}")
        return str(filepath), filename
    
    def generate_outstanding_report(self, report_type: str = "receivable") -> Tuple[str, str]:
        """
        Generate Outstanding Receivables or Payables Excel report.
        
        Args:
            report_type: "receivable" or "payable"
        """
        if report_type == "receivable":
            group_filter = "%sundry debtor%"
            title = "Outstanding Receivables"
        else:
            group_filter = "%sundry creditor%"
            title = "Outstanding Payables"
        
        # Use 'parent' field which is the actual field name in Ledger model
        ledgers = self.db.query(Ledger).filter(
            Ledger.parent.ilike(group_filter),
            Ledger.closing_balance != 0,
            Ledger.tenant_id == self.tenant_id
        ).order_by(desc(Ledger.closing_balance)).all()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel has 31 char limit
        
        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = f"{title} as on {datetime.now().strftime('%d-%b-%Y')}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Company name
        ws['A2'] = "KRISHA SALES"
        ws['A2'].font = Font(bold=True, size=10, color="666666")
        
        # Headers
        headers = ["#", "Party Name", "Group", "Opening Balance", "Closing Balance"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Data
        total_opening = 0.0
        total_closing = 0.0
        
        for row_idx, ledger in enumerate(ledgers, 5):
            ws.cell(row=row_idx, column=1, value=row_idx - 4)
            ws.cell(row=row_idx, column=2, value=ledger.name)
            ws.cell(row=row_idx, column=3, value=ledger.parent or "")
            ws.cell(row=row_idx, column=4, value=ledger.opening_balance or 0).number_format = '₹#,##0.00'
            ws.cell(row=row_idx, column=5, value=ledger.closing_balance or 0).number_format = '₹#,##0.00'
            
            total_opening += ledger.opening_balance or 0
            total_closing += ledger.closing_balance or 0
            
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).border = self.border
        
        # Total
        total_row = len(ledgers) + 5
        ws.cell(row=total_row, column=2, value="TOTAL:").font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=total_opening).number_format = '₹#,##0.00'
        ws.cell(row=total_row, column=4).font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=total_closing).number_format = '₹#,##0.00'
        ws.cell(row=total_row, column=5).font = Font(bold=True)
        
        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        
        filename = f"{title.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = self.exports_dir / filename
        wb.save(filepath)
        
        logger.info(f"Generated outstanding report Excel: {filepath}")
        return str(filepath), filename
    
    def generate_stock_report(self) -> Tuple[str, str]:
        """Generate Stock/Inventory Excel report."""
        items = self.db.query(StockItem).filter(
            StockItem.tenant_id == self.tenant_id
        ).order_by(StockItem.name).all()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock Report"
        
        # Title
        ws.merge_cells('A1:G1')
        ws['A1'] = f"Stock Report as on {datetime.now().strftime('%d-%b-%Y')}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Company name
        ws['A2'] = "KRISHA SALES"
        ws['A2'].font = Font(bold=True, size=10, color="666666")
        
        # Headers - use actual StockItem fields
        headers = ["#", "Item Name", "Unit", "Opening Qty", "Closing Qty", "Rate", "HSN"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Data - StockItem fields: units, opening_stock, closing_balance, selling_price, hsn_code
        for row_idx, item in enumerate(items, 5):
            ws.cell(row=row_idx, column=1, value=row_idx - 4)
            ws.cell(row=row_idx, column=2, value=item.name)
            ws.cell(row=row_idx, column=3, value=item.units or "Nos")
            ws.cell(row=row_idx, column=4, value=item.opening_stock or 0)
            ws.cell(row=row_idx, column=5, value=item.closing_balance or 0)
            ws.cell(row=row_idx, column=6, value=item.selling_price or item.rate or 0).number_format = '₹#,##0.00'
            ws.cell(row=row_idx, column=7, value=item.hsn_code or "-")
            
            for col in range(1, 8):
                ws.cell(row=row_idx, column=col).border = self.border
        
        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        
        filename = f"Stock_Report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = self.exports_dir / filename
        wb.save(filepath)
        
        logger.info(f"Generated stock report Excel: {filepath}")
        return str(filepath), filename


class ExportService:
    """
    Main Export Service - Unified interface for all export types.
    Used by the Query Orchestrator when user requests PDF/Excel.
    """
    
    def __init__(self, db: Session, tenant_id: str = "default"):
        self.db = db
        self.tenant_id = tenant_id
        self.pdf = PDFGenerator(db, tenant_id)
        self.excel = ExcelGenerator(db, tenant_id)
    
    def export_invoice_pdf(self, voucher_id: int) -> Dict[str, Any]:
        """Export a single invoice as PDF"""
        try:
            filepath, filename = self.pdf.generate_invoice_pdf(voucher_id)
            return {
                "success": True,
                "file_path": filepath,
                "filename": filename,
                "type": "pdf",
                "message": f"📄 Invoice PDF generated: {filename}"
            }
        except Exception as e:
            logger.error(f"PDF export failed: {e}")
            return {"success": False, "error": str(e)}
    
    def export_statement_pdf(self, party_name: str) -> Dict[str, Any]:
        """Export outstanding statement as PDF"""
        try:
            filepath, filename = self.pdf.generate_outstanding_statement(party_name)
            return {
                "success": True,
                "file_path": filepath,
                "filename": filename,
                "type": "pdf",
                "message": f"📄 Statement PDF generated: {filename}"
            }
        except Exception as e:
            logger.error(f"Statement PDF export failed: {e}")
            return {"success": False, "error": str(e)}
    
    def export_sales_excel(
        self, 
        date_from: datetime = None, 
        date_to: datetime = None
    ) -> Dict[str, Any]:
        """Export sales register as Excel"""
        try:
            filepath, filename = self.excel.generate_sales_register(date_from, date_to)
            return {
                "success": True,
                "file_path": filepath,
                "filename": filename,
                "type": "excel",
                "message": f"📊 Sales Register Excel generated: {filename}"
            }
        except Exception as e:
            logger.error(f"Sales Excel export failed: {e}")
            return {"success": False, "error": str(e)}
    
    def export_purchase_excel(
        self, 
        date_from: datetime = None, 
        date_to: datetime = None
    ) -> Dict[str, Any]:
        """Export purchase register as Excel"""
        try:
            filepath, filename = self.excel.generate_purchase_register(date_from, date_to)
            return {
                "success": True,
                "file_path": filepath,
                "filename": filename,
                "type": "excel",
                "message": f"📊 Purchase Register Excel generated: {filename}"
            }
        except Exception as e:
            logger.error(f"Purchase Excel export failed: {e}")
            return {"success": False, "error": str(e)}
    
    def export_outstanding_excel(self, report_type: str = "receivable") -> Dict[str, Any]:
        """Export outstanding report as Excel"""
        try:
            filepath, filename = self.excel.generate_outstanding_report(report_type)
            return {
                "success": True,
                "file_path": filepath,
                "filename": filename,
                "type": "excel",
                "message": f"📊 Outstanding Report Excel generated: {filename}"
            }
        except Exception as e:
            logger.error(f"Outstanding Excel export failed: {e}")
            return {"success": False, "error": str(e)}
    
    def export_stock_excel(self) -> Dict[str, Any]:
        """Export stock report as Excel"""
        try:
            filepath, filename = self.excel.generate_stock_report()
            return {
                "success": True,
                "file_path": filepath,
                "filename": filename,
                "type": "excel",
                "message": f"📊 Stock Report Excel generated: {filename}"
            }
        except Exception as e:
            logger.error(f"Stock Excel export failed: {e}")
            return {"success": False, "error": str(e)}


# ============== CONVENIENCE FUNCTIONS ==============

def export_invoice_to_pdf(voucher_id: int, tenant_id: str = "default") -> Dict[str, Any]:
    """Quick function to export invoice PDF"""
    db = SessionLocal()
    try:
        service = ExportService(db, tenant_id)
        return service.export_invoice_pdf(voucher_id)
    finally:
        db.close()


def export_statement_to_pdf(party_name: str, tenant_id: str = "default") -> Dict[str, Any]:
    """Quick function to export statement PDF"""
    db = SessionLocal()
    try:
        service = ExportService(db, tenant_id)
        return service.export_statement_pdf(party_name)
    finally:
        db.close()


def export_sales_to_excel(
    date_from: datetime = None, 
    date_to: datetime = None, 
    tenant_id: str = "default"
) -> Dict[str, Any]:
    """Quick function to export sales Excel"""
    db = SessionLocal()
    try:
        service = ExportService(db, tenant_id)
        return service.export_sales_excel(date_from, date_to)
    finally:
        db.close()
