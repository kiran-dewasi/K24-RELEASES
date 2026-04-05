"""
utils/excel_generator.py

Standalone Excel workbook generator for K24 reports.
No FastAPI / SQLAlchemy / database imports — pure openpyxl.
Called by routers/reports.py export-excel endpoint.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def generate_report_excel(
    slug: str,
    report_title: str,
    data: dict,
    period: str,
) -> bytes:
    """
    Generate a minimal valid XLSX workbook for the given report slug.

    Structure:
        Row 1: Report title
        Row 2: Period
        Row 3: Generated timestamp
        Row 4: (blank)
        Row 5: Column headers (blue background, white bold text)
        Row 6+: Data rows

    Returns raw bytes suitable for streaming as an HTTP response.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_title[:31]  # Excel sheet name max 31 chars

    # ── Styles ──────────────────────────────────────────────────────────────
    header_fill  = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font  = Font(bold=True, color="FFFFFF", size=10)
    bold_font    = Font(bold=True, size=10)
    center_align = Alignment(horizontal="center")

    # ── Meta rows ────────────────────────────────────────────────────────────
    ws.append([f"Report: {report_title}"])
    ws["A1"].font = bold_font
    ws.append([f"Period: {period}"])
    ws.append([f"Generated: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}"])
    ws.append([])  # blank spacer — row 4

    # ── Headers + rows by slug ───────────────────────────────────────────────
    if slug in ("sales-register", "purchase-register",
                "receipt-register", "payment-register",
                "journal-register", "cash-flow"):
        headers = ["#", "Date", "Party Name", "Voucher Type",
                   "Voucher No.", "Narration", "Amount (INR)"]
        vouchers = data.get("vouchers", [])
        rows = [
            [
                i + 1,
                v.get("date", ""),
                v.get("party_name", ""),
                v.get("voucher_type", "") or v.get("direction", ""),
                v.get("voucher_no", ""),
                v.get("narration", ""),
                v.get("amount", 0),
            ]
            for i, v in enumerate(vouchers)
        ]

    elif slug == "balance-sheet":
        headers = ["#", "Ledger Name", "Group", "Category", "Amount (INR)"]
        rows = []
        for i, item in enumerate(data.get("assets", []), 1):
            rows.append([i, item.get("name", ""), item.get("group", ""),
                         "Asset", item.get("amount", 0)])
        for item in data.get("liabilities", []):
            rows.append([len(rows) + 1, item.get("name", ""), item.get("group", ""),
                         "Liability", item.get("amount", 0)])

    elif slug == "profit-loss":
        headers = ["#", "Account Head", "Category", "Amount (INR)"]
        rows = []
        for k, v in data.get("income", {}).items():
            rows.append([len(rows) + 1, k, "Income", v])
        for k, v in data.get("expenses", {}).items():
            rows.append([len(rows) + 1, k, "Expense", v])

    else:
        # Generic fallback — always produces at least one data row
        headers = ["Report Type", "Note"]
        rows    = [[slug, "TEMPORARY EXPORT PLACEHOLDER"]]

    # ── Write header row (row 5) ─────────────────────────────────────────────
    ws.append(headers)
    hdr_row = ws.max_row
    for cell in ws[hdr_row]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align

    # ── Write data rows ──────────────────────────────────────────────────────
    for row in rows:
        ws.append(row)

    # ── Auto-size columns (rough estimate) ───────────────────────────────────
    for col_cells in ws.columns:
        max_len = max(
            (len(str(cell.value or "")) for cell in col_cells),
            default=8,
        )
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
